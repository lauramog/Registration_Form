import openpyxl
from tkinter import *
from tkinter import messagebox

##################################################
## Lectura de libro de excel, hojas y columnas ##
##################################################

path='elprogram.xlsx'
wb= openpyxl.load_workbook(path)
	
sheet1=wb['MateriasyCupos']
sheet2=wb['EstudiantesMatriculados']

Sub=[]
Cupos=[]

multiple_cells1 = sheet1['A1':'A8']
for row in multiple_cells1:
    for cell in row:
        Sub.append(cell.value)


multiple_cells2 = sheet1['B1':'B8']
for row in multiple_cells2:
    for cell in row:
        Cupos.append(cell.value)

########################################################################################
##  definiendo la funcion que guarda en la hoja de excel el nombre de los Estudiantes ##
########################################################################################
def clicked():

	name=(name_entry.get())


	for i in range(2,20):
		if sheet2.cell(row = i, column=1).value == None:
			sheet2.cell(row=i, column=1).value = name	
		else:
			continue 
		break


	wb.save(path)
	name_entry.delete(0,END)
	Accept.configure(text= "Aceptar")

###############################################################################################
## definiendo la funcion que guarda en la hoja de excel el nombre de la materia seleccionada ##
###############################################################################################

def get_list():
	x= Subj.curselection()
	y= Subj.get(x)
	print(x)
	
	for i in range(2,20):
		if sheet2.cell(row = i, column=2).value == None:
			sheet2.cell(row=i, column=2).value = y	
		else:
			continue 
		break


	if sheet1.cell(row = 1, column=2).value==-1:
		messagebox.showinfo('Limite de cupos excedido', 'No quedan cupos disponible en esta materia')


	if x== (0,):
		sheet1.cell(row = 1, column=2).value = sheet1.cell(row = 1, column=2).value - 1

	if x== (1,):
		sheet1.cell(row = 2, column=2).value = sheet1.cell(row = 1, column=2).value - 1

	if x== (2,):
		sheet1.cell(row = 3, column=2).value = sheet1.cell(row = 1, column=2).value - 1

	if x== (3,):
		sheet1.cell(row = 4, column=2).value = sheet1.cell(row = 1, column=2).value - 1

	if x== (4,):
		sheet1.cell(row = 5, column=2).value = sheet1.cell(row = 1, column=2).value - 1
	
	if x== (5,):
		sheet1.cell(row = 6, column=2).value = sheet1.cell(row = 1, column=2).value - 1



	   	



	wb.save(path)
 						
						##############
						## Interfaz ##
						##############

window = Tk()
window.title("Universidad de Antioquia")
window.geometry('450x400')

#########################
## Seleccion de Materia##
#########################

Subj=Label(window,text='Seleccione una materia:',font=("Arial Bold", 15)).place(x=1,y=15)
Subj = Listbox(window,height=8,width=10,selectmode=BROWSE)
Subj.pack() 
Subj.insert(END,*Sub)


Select = Button(window, text = "Seleccionar", bg="grey", command= get_list)
Select.place(x=185 , y= 140)

########################
## Nombre del usuario ##
########################

name = 	StringVar()
name_text=Label(text= "Ingrese su nombre ",font=("Arial Bold", 15))
name_text.place(x=1, y=250)
name_entry = Entry(window, width=40)	
name_entry.place(x=1, y= 270)


Accept = Button(window, text = "Aceptar", bg="grey", command= clicked)
Accept.place(x= 150, y= 320)

window.mainloop()

