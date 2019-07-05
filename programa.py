import numpy as np
import pandas as pd 
from openpyxl import load_workbook
from tkinter import *

path = 'program.xlsx'
df1 =   pd.read_excel(path)
SeriesA=df1['MATERIAS']
def clicked():
	path = 'program.xlsx'
	df1   =   pd.read_excel(path,sheet_name='EstudiantesMatriculados')	
	book=load_workbook(path)
	
	name    =  df1['NOMBRE DEL ESTUDIANTE']
	subject =  df1['MATERIA']
	

	name=(name_entry.get())
	sub=(sub_entry.get())
	
	#name= name.append(a)
	#sub= sub.append(b)

	df=pd.DataFrame({"name":name,"subject":sub})
	df.to_excel(path,index=False)	

	writer = pd.ExcelWriter('program.xlsx', sheet_name='EstudiantesMatriculados',engine= 'openpyxl')
	writer.book=writer
	
	#df1.to_excel(writer,sheet_name='EstudiantesMatriculados', index=False)
	

	
	writer.save()
	#writer.close()

	name_entry.delete(0,END)
	sub_entry.delete(0,END)
	
	
	
	Accept.configure(text= "ha quedado Matriculado")
	



window = Tk()
window.title("Universidad de Antioquia")
window.geometry('450x400')
lbl = Label(window, text=SeriesA[0],font=("Arial Bold", 15))
lbl.grid(column=0, row=0)

btn = Button(window, text="Seleccionar", bg="white", fg="black")
btn.grid(column=1, row=0)

name = 	StringVar()
name_text=Label(text= "Ingrese su nombre ",font=("Arial Bold", 15))
name_text.place(x=1, y=50)
name_entry = Entry(window, width=40)	
name_entry.place(x=1, y= 80)



	
sub_text=Label(text= "Nombre de la materia ",font=("Arial Bold", 15))
sub_text.place(x=1, y=120)
sub_entry = Entry(window, width=40)
sub_entry.place(x=1, y= 150)



Accept = Button(window, text = "Aceptar", bg="grey", command= clicked)
Accept.place(x= 150, y= 190)

window.mainloop()


#writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#book= load_workbook(path)
#writer.book=book
